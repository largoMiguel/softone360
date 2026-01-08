"""
Servicio para generar reportes en Excel del PDM
"""
from io import BytesIO
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.models.pdm import PdmProducto
from app.models.secretaria import Secretaria


class PDMExcelGenerator:
    """Generador de reportes Excel para PDM"""
    
    # Estilos
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    @staticmethod
    def generar_plan_accion(
        db: Session,
        entity_id: int,
        anio: int,
        secretaria_ids: Optional[List[int]] = None
    ) -> BytesIO:
        """
        Genera un Excel con el Plan de Acción del PDM
        
        Args:
            db: Sesión de base de datos
            entity_id: ID de la entidad
            anio: Año del plan (2024-2027)
            secretaria_ids: Lista de IDs de secretarías (opcional, para filtrar)
            
        Returns:
            BytesIO con el archivo Excel
        """
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Plan Acción {anio}"
        
        # Configurar encabezados
        headers = [
            "ITEM",
            "LÍNEA ESTRATÉGICA",
            "SECTOR",
            "PROGRAMA MGA",
            "CÓDIGO DE PRODUCTO",
            "PRODUCTO",
            "INDICADOR DE PRODUCTO",
            f"CANTIDAD META PROGRAMADA AÑO {anio}",
            "UNIDAD DE MEDIDA",
            f"PRESUPUESTO ASIGNADO {anio} (Según Ejecución)",
            "DEPENDENCIA RESPONSABLE"
        ]
        
        # Escribir encabezados
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = PDMExcelGenerator.HEADER_FONT
            cell.fill = PDMExcelGenerator.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = PDMExcelGenerator.BORDER
        
        # Query de productos
        query = db.query(PdmProducto).filter(
            PdmProducto.entity_id == entity_id
        )
        
        # Filtrar por secretarías si se especifican
        if secretaria_ids and len(secretaria_ids) > 0:
            query = query.filter(PdmProducto.responsable_secretaria_id.in_(secretaria_ids))
        
        # Ordenar por línea estratégica y código de producto
        query = query.order_by(
            PdmProducto.linea_estrategica,
            PdmProducto.codigo_producto
        )
        
        productos = query.all()
        
        # Escribir datos
        for idx, producto in enumerate(productos, 2):
            # Obtener meta programada según el año
            meta_programada = 0
            presupuesto_asignado = 0
            
            if anio == 2024:
                meta_programada = producto.programacion_2024 or 0
                presupuesto_asignado = producto.total_2024 or 0
            elif anio == 2025:
                meta_programada = producto.programacion_2025 or 0
                presupuesto_asignado = producto.total_2025 or 0
            elif anio == 2026:
                meta_programada = producto.programacion_2026 or 0
                presupuesto_asignado = producto.total_2026 or 0
            elif anio == 2027:
                meta_programada = producto.programacion_2027 or 0
                presupuesto_asignado = producto.total_2027 or 0
            
            # Obtener nombre de secretaría
            dependencia = producto.responsable_secretaria_nombre or ""
            if producto.responsable_secretaria:
                dependencia = producto.responsable_secretaria.nombre
            
            # Escribir fila
            row_data = [
                idx - 1,  # ITEM
                producto.linea_estrategica or "",
                producto.sector_mga or "",
                producto.programa_mga or "",
                producto.codigo_producto or "",
                producto.producto_mga or "",
                producto.indicador_producto_mga or "",
                meta_programada,
                producto.unidad_medida or "",
                presupuesto_asignado,
                dependencia
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=idx, column=col, value=value)
                cell.border = PDMExcelGenerator.BORDER
                
                # Formato para números
                if col in [8, 10]:  # Meta y Presupuesto
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Ajustar anchos de columna
        column_widths = {
            'A': 8,   # ITEM
            'B': 35,  # LÍNEA ESTRATÉGICA
            'C': 25,  # SECTOR
            'D': 40,  # PROGRAMA MGA
            'E': 18,  # CÓDIGO DE PRODUCTO
            'F': 45,  # PRODUCTO
            'G': 45,  # INDICADOR
            'H': 18,  # META
            'I': 20,  # UNIDAD
            'J': 20,  # PRESUPUESTO
            'K': 30   # DEPENDENCIA
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Congelar primera fila
        ws.freeze_panes = 'A2'
        
        # Guardar en BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
